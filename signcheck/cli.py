import csv
import json
import os
import sys
import zipfile
import tempfile
import shutil
import click
from datetime import datetime
from typing import List, Optional, Dict, Any
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .models import (
    EnrollmentRecord,
    SigninRecord,
    MatchRule,
    FieldMapping,
    ImportErrorRecord,
)
from .storage import Storage
from .reconcile import reconcile, undo, mark_result, batch_mark
from . import config as config_module


VALID_STATUSES = {"normal", "absent", "non_enrolled", "duplicate"}

STATUS_LABELS = {
    "normal": "正常签到",
    "absent": "缺席",
    "non_enrolled": "非报名人员",
    "duplicate": "重复扫码",
}


def _write_json(file_path: str, data: Any):
    abs_path = os.path.abspath(file_path)
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)
    with open(abs_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _write_xlsx(file_path: str, headers: List[str], rows: List[List[Any]], sheet_name: str = "Sheet1"):
    if not HAS_OPENPYXL:
        raise RuntimeError("openpyxl 未安装，请执行：pip install openpyxl>=3.1.0")
    abs_path = os.path.abspath(file_path)
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font
        cell.alignment = center_align

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"

    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    wb.save(abs_path)


def _resolve_filters(
    status, session, mark, keyword, limit, sort_by, sort_order, view, save_view, overwrite, storage
) -> Dict[str, Any]:
    filters: Dict[str, Any] = {}
    if view:
        saved = storage.get_view(view)
        if saved is None:
            click.echo(f"错误：视图「{view}」不存在", err=True)
            storage.close()
            sys.exit(1)
        view_filters = json.loads(saved["filters"])
        filters.update(view_filters)
    cli_overrides: Dict[str, Any] = {}
    if status is not None:
        cli_overrides["status"] = status
    if session is not None:
        cli_overrides["session"] = session
    if mark is not None:
        cli_overrides["mark"] = mark
    if keyword is not None:
        cli_overrides["keyword"] = keyword
    if limit is not None:
        cli_overrides["limit"] = limit
    if sort_by is not None:
        cli_overrides["sort_by"] = sort_by
    if sort_order is not None:
        cli_overrides["sort_order"] = sort_order
    filters.update(cli_overrides)
    if "sort_by" not in filters or filters["sort_by"] is None:
        filters["sort_by"] = config_module.get_config("sort_by")
    if "sort_order" not in filters or filters["sort_order"] is None:
        filters["sort_order"] = config_module.get_config("sort_order")
    if save_view:
        filters_to_save = {k: v for k, v in filters.items() if v is not None}
        saved_json = json.dumps(filters_to_save, ensure_ascii=False)
        ok = storage.save_view(save_view, saved_json, overwrite=overwrite)
        if not ok:
            click.echo(f"错误：视图「{save_view}」已存在，使用 --overwrite 覆盖", err=True)
            storage.close()
            sys.exit(1)
        click.echo(f"[OK] 视图「{save_view}」已保存")
    return filters


def _validate_status(status: Optional[str]) -> Optional[str]:
    if status is None:
        return None
    if status not in VALID_STATUSES:
        click.echo(f"错误：非法状态「{status}」，可选值：{', '.join(sorted(VALID_STATUSES))}", err=True)
        sys.exit(1)
    return status


def _get_storage() -> Storage:
    return Storage()


def _read_csv(file_path: str) -> tuple:
    abs_path = os.path.abspath(file_path)
    if not os.path.exists(abs_path):
        click.echo(f"错误：文件不存在 {abs_path}", err=True)
        sys.exit(1)
    with open(abs_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    return rows, abs_path


def _apply_field_mapping(rows: list, mapping: dict) -> list:
    if not mapping:
        return rows
    mapped = []
    for row in rows:
        new_row = {}
        for field_name, csv_col in mapping.items():
            new_row[field_name] = row.get(csv_col, "").strip()
        mapped.append(new_row)
    return mapped


@click.group()
@click.version_option(version="1.0.0")
def main():
    """线下培训签到对账 CLI"""
    pass


@main.command("import-enroll")
@click.argument("csv_file")
@click.option("--dry-run", is_flag=True, default=False, help="只校验不落库，预览导入结果")
def import_enroll(csv_file: str, dry_run: bool):
    """导入报名 CSV 文件"""
    storage = _get_storage()
    rows, abs_path = _read_csv(csv_file)
    field_mapping = storage.get_field_mapping()
    mapping = field_mapping.enroll if field_mapping.enroll else {
        "name": "姓名",
        "phone": "手机号",
        "session": "场次",
    }

    mapped_rows = _apply_field_mapping(rows, mapping)

    valid_records: List[EnrollmentRecord] = []
    errors: List[ImportErrorRecord] = []
    existing_sessions = storage.get_enrollment_sessions()

    for idx, row in enumerate(mapped_rows):
        source_row = idx + 2
        name = row.get("name", "").strip()
        phone = row.get("phone", "").strip() if row.get("phone") else ""
        session = row.get("session", "").strip()

        row_errors = []
        if not phone:
            row_errors.append(("missing_phone", f"第 {source_row} 行：手机号缺失"))
        if not session:
            row_errors.append(("missing_session", f"第 {source_row} 行：场次不能为空"))
        if session and storage.is_session_closed(session):
            row_errors.append(("session_closed", f"第 {source_row} 行：场次「{session}」已关闭，无法导入"))

        if row_errors:
            for error_type, error_msg in row_errors:
                raw = json.dumps({k: v for k, v in row.items() if v}, ensure_ascii=False)
                errors.append(ImportErrorRecord(
                    source_type="enroll",
                    source_file=os.path.basename(abs_path),
                    row_number=source_row,
                    error_type=error_type,
                    error_message=error_msg,
                    raw_data=raw,
                ))
            continue

        valid_records.append(EnrollmentRecord(
            name=name,
            phone=phone,
            session=session,
            source_file=os.path.basename(abs_path),
            source_row=source_row,
        ))

    if dry_run:
        click.echo(f"[DRY-RUN] 校验完成，将导入 {len(valid_records)} 条报名记录")
        skipped = len(errors)
        if skipped:
            click.echo(f"[DRY-RUN] 跳过 {skipped} 条问题记录：")
            for e in errors:
                click.echo(f"  [ERR] {e.error_message}")
        else:
            click.echo("[DRY-RUN] 所有记录校验通过，无错误")
        storage.close()
        return

    if valid_records:
        new_ids, existing_ids = storage.add_enrollments(valid_records)
        if new_ids:
            undo_data = json.dumps({"action": "import_enroll", "ids": new_ids})
            storage.add_undo_action("import_enroll", undo_data)
    else:
        new_ids, existing_ids = [], []

    if errors:
        storage.add_import_errors(errors)

    click.echo(f"[OK] 新增 {len(new_ids)} 条报名记录，跳过 {len(existing_ids)} 条已存在记录")
    if errors:
        for e in errors:
            click.echo(f"[ERR] {e.error_message}")
        click.echo(f"共 {len(errors)} 条错误，已跳过")

    storage.close()


@main.command("import-signin")
@click.argument("csv_file")
@click.option("--dry-run", is_flag=True, default=False, help="只校验不落库，预览导入结果")
def import_signin(csv_file: str, dry_run: bool):
    """导入扫码签到 CSV 文件"""
    storage = _get_storage()
    rows, abs_path = _read_csv(csv_file)
    field_mapping = storage.get_field_mapping()
    mapping = field_mapping.signin if field_mapping.signin else {
        "name": "姓名",
        "phone": "手机号",
        "session": "场次",
        "scan_time": "扫码时间",
    }

    mapped_rows = _apply_field_mapping(rows, mapping)

    valid_records: List[SigninRecord] = []
    errors: List[ImportErrorRecord] = []
    existing_sessions = storage.get_enrollment_sessions()

    for idx, row in enumerate(mapped_rows):
        source_row = idx + 2
        name = row.get("name", "").strip()
        phone = row.get("phone", "").strip() if row.get("phone") else ""
        session = row.get("session", "").strip()
        scan_time = row.get("scan_time", "").strip() if row.get("scan_time") else ""

        row_errors = []
        if not phone:
            row_errors.append(("missing_phone", f"第 {source_row} 行：手机号缺失"))
        if existing_sessions and session and session not in existing_sessions:
            row_errors.append(("invalid_session", f"第 {source_row} 行：场次「{session}」不存在"))
        if session and storage.is_session_closed(session):
            row_errors.append(("session_closed", f"第 {source_row} 行：场次「{session}」已关闭，无法导入"))

        if row_errors:
            for error_type, error_msg in row_errors:
                raw = json.dumps({k: v for k, v in row.items() if v}, ensure_ascii=False)
                errors.append(ImportErrorRecord(
                    source_type="signin",
                    source_file=os.path.basename(abs_path),
                    row_number=source_row,
                    error_type=error_type,
                    error_message=error_msg,
                    raw_data=raw,
                ))
            continue

        valid_records.append(SigninRecord(
            name=name,
            phone=phone,
            session=session,
            scan_time=scan_time,
            source_file=os.path.basename(abs_path),
            source_row=source_row,
        ))

    if dry_run:
        click.echo(f"[DRY-RUN] 校验完成，将导入 {len(valid_records)} 条签到记录")
        skipped = len(errors)
        if skipped:
            click.echo(f"[DRY-RUN] 跳过 {skipped} 条问题记录：")
            for e in errors:
                click.echo(f"  [ERR] {e.error_message}")
        else:
            click.echo("[DRY-RUN] 所有记录校验通过，无错误")
        storage.close()
        return

    if valid_records:
        new_ids, existing_ids = storage.add_signins(valid_records)
        if new_ids:
            undo_data = json.dumps({"action": "import_signin", "ids": new_ids})
            storage.add_undo_action("import_signin", undo_data)
    else:
        new_ids, existing_ids = [], []

    if errors:
        storage.add_import_errors(errors)

    click.echo(f"[OK] 新增 {len(new_ids)} 条签到记录，跳过 {len(existing_ids)} 条已存在记录")
    if errors:
        for e in errors:
            click.echo(f"[ERR] {e.error_message}")
        click.echo(f"共 {len(errors)} 条错误，已跳过")

    storage.close()


@main.command("import-rules")
@click.argument("json_file")
@click.option("--dry-run", is_flag=True, default=False, help="只校验不落库，预览导入结果")
def import_rules(json_file: str, dry_run: bool):
    """导入匹配规则 JSON 文件"""
    storage = _get_storage()
    abs_path = os.path.abspath(json_file)
    if not os.path.exists(abs_path):
        click.echo(f"错误：文件不存在 {abs_path}", err=True)
        sys.exit(1)

    with open(abs_path, "r", encoding="utf-8-sig") as f:
        data = json.load(f)

    match_rules_data = data.get("match_rules", [])
    rules: List[MatchRule] = []
    errors: List[str] = []

    for idx, r in enumerate(match_rules_data):
        rule_idx = idx + 1
        if "field" not in r:
            errors.append(f"第 {rule_idx} 条规则：缺少 field 字段")
            continue
        field_name = r["field"]
        match_type = r.get("match_type", "exact")
        if match_type not in ("exact", "fuzzy"):
            errors.append(f"第 {rule_idx} 条规则：match_type「{match_type}」不合法，可选 exact/fuzzy")
            continue
        threshold = r.get("threshold")
        if match_type == "fuzzy" and threshold is None:
            errors.append(f"第 {rule_idx} 条规则：fuzzy 匹配必须指定 threshold")
            continue
        rules.append(MatchRule(
            field_name=field_name,
            match_type=match_type,
            threshold=threshold,
            priority=r.get("priority", 0),
        ))

    field_mapping_data = data.get("field_mapping", None)
    if field_mapping_data:
        if not isinstance(field_mapping_data, dict):
            errors.append("field_mapping 必须是对象")
        else:
            for key in ("enroll", "signin"):
                if key in field_mapping_data and not isinstance(field_mapping_data[key], dict):
                    errors.append(f"field_mapping.{key} 必须是对象")

    if dry_run:
        click.echo(f"[DRY-RUN] 校验完成，将导入 {len(rules)} 条匹配规则")
        if field_mapping_data:
            click.echo("[DRY-RUN] 将更新字段映射配置")
        if errors:
            click.echo(f"[DRY-RUN] 发现 {len(errors)} 个错误：")
            for e in errors:
                click.echo(f"  [ERR] {e}")
        else:
            click.echo("[DRY-RUN] 所有规则校验通过，无错误")
        storage.close()
        return

    if errors:
        click.echo(f"错误：校验未通过，共 {len(errors)} 个问题：", err=True)
        for e in errors:
            click.echo(f"  [ERR] {e}", err=True)
        storage.close()
        sys.exit(1)

    prev_rules = storage.get_all_rules()
    prev_mapping = storage.get_field_mapping()

    storage.clear_rules()

    if rules:
        storage.add_rules(rules)

    if field_mapping_data:
        fm = FieldMapping(
            enroll=field_mapping_data.get("enroll", {}),
            signin=field_mapping_data.get("signin", {}),
        )
        storage.save_field_mapping(fm)

    undo_data = json.dumps({
        "action": "import_rules",
        "prev_rules": [{"field_name": r.field_name, "match_type": r.match_type, "threshold": r.threshold, "priority": r.priority} for r in prev_rules],
        "prev_mapping": {"enroll": prev_mapping.enroll, "signin": prev_mapping.signin} if prev_mapping.enroll or prev_mapping.signin else None,
    }, ensure_ascii=False)
    storage.add_undo_action("import_rules", undo_data)

    click.echo(f"[OK] 成功导入 {len(rules)} 条匹配规则")
    if field_mapping_data:
        click.echo(f"[OK] 成功导入字段映射配置")

    storage.close()


@main.command("reconcile")
def do_reconcile():
    """执行对账"""
    storage = _get_storage()
    enrollments = storage.get_all_enrollments()
    signins = storage.get_all_signins()

    if not enrollments:
        click.echo("错误：尚未导入报名数据，请先执行 import-enroll", err=True)
        storage.close()
        sys.exit(1)
    if not signins:
        click.echo("错误：尚未导入签到数据，请先执行 import-signin", err=True)
        storage.close()
        sys.exit(1)

    results, skipped_sessions = reconcile(storage)

    if skipped_sessions:
        click.echo(f"[提示] 已跳过 {len(skipped_sessions)} 个已关闭场次：{', '.join(skipped_sessions)}")

    status_counts = {}
    for r in results:
        status_counts[r.status] = status_counts.get(r.status, 0) + 1

    click.echo("对账完成：")
    for status in ["normal", "absent", "non_enrolled", "duplicate"]:
        cnt = status_counts.get(status, 0)
        label = STATUS_LABELS.get(status, status)
        click.echo(f"  {label}: {cnt}")

    storage.close()


@main.command("mark")
@click.argument("result_id", type=int)
@click.option("--mark-text", "-m", default=None, help="标记内容")
@click.option("--notes", "-n", default=None, help="备注")
def do_mark(result_id: int, mark_text: Optional[str], notes: Optional[str]):
    """人工标记对账结果"""
    storage = _get_storage()

    if mark_text is None and notes is None:
        click.echo("错误：请至少提供 --mark-text 或 --notes", err=True)
        storage.close()
        sys.exit(1)

    msg = mark_result(storage, result_id, mark_text, notes)
    if msg is None:
        click.echo(f"错误：未找到结果 #{result_id}", err=True)
        storage.close()
        sys.exit(1)

    click.echo(f"[OK] {msg}")
    storage.close()


@main.command("undo")
def do_undo():
    """撤销上一步操作"""
    storage = _get_storage()
    msg = undo(storage)
    if msg is None:
        click.echo("没有可撤销的操作")
    else:
        click.echo(f"[OK] 已撤销上一步操作：{msg}")
    storage.close()


def _common_filter_options(f):
    f = click.option("--status", "-s", default=None, help="按状态筛选 (normal/absent/non_enrolled/duplicate)")(f)
    f = click.option("--session", default=None, help="按场次筛选")(f)
    f = click.option("--mark", default=None, help="按人工标记筛选")(f)
    f = click.option("--keyword", "-k", default=None, help="按姓名或手机号关键词筛选")(f)
    f = click.option("--limit", "-n", type=int, default=None, help="限制返回条数")(f)
    f = click.option("--sort-by", type=click.Choice(["id", "status"]), default=None, help="排序字段（默认 id）")(f)
    f = click.option("--sort-order", type=click.Choice(["asc", "desc"]), default=None, help="排序方向（默认 asc）")(f)
    f = click.option("--view", "-v", default=None, help="使用已保存的视图名加载筛选条件")(f)
    f = click.option("--save-view", default=None, help="将当前筛选条件保存为命名视图")(f)
    f = click.option("--overwrite", is_flag=True, default=False, help="覆盖同名视图")(f)
    return f


@main.command("list")
@_common_filter_options
def do_list(status, session, mark, keyword, limit, sort_by, sort_order, view, save_view, overwrite):
    """查看对账结果列表（支持筛选、排序、视图）"""
    storage = _get_storage()
    if status is not None:
        status = _validate_status(status)
    filters = _resolve_filters(
        status, session, mark, keyword, limit, sort_by, sort_order, view, save_view, overwrite, storage
    )
    results = storage.query_reconcile_results(
        status=filters.get("status"),
        session=filters.get("session"),
        mark=filters.get("mark"),
        keyword=filters.get("keyword"),
        limit=filters.get("limit"),
        sort_by=filters.get("sort_by", "id"),
        sort_order=filters.get("sort_order", "asc"),
    )

    if not results:
        click.echo("暂无匹配的对账结果")
        storage.close()
        return

    col_widths = {
        "id": max(4, max(len(str(r.id)) for r in results)),
        "name": max(4, max(len(r.name) for r in results)),
        "phone": max(4, max(len(r.phone or "") for r in results)),
        "session": max(4, max(len(r.session) for r in results)),
        "status": max(4, max(len(STATUS_LABELS.get(r.status, r.status)) for r in results)),
        "mark": max(4, max(len(r.manual_mark or "-") for r in results)),
        "notes": max(4, max(len(r.notes or "-") for r in results)),
    }

    header = (
        f"{'ID':<{col_widths['id']}}  "
        f"{'姓名':<{col_widths['name']}}  "
        f"{'手机号':<{col_widths['phone']}}  "
        f"{'场次':<{col_widths['session']}}  "
        f"{'状态':<{col_widths['status']}}  "
        f"{'标记':<{col_widths['mark']}}  "
        f"{'备注':<{col_widths['notes']}}"
    )
    click.echo(header)
    click.echo("-" * len(header))

    for r in results:
        label = STATUS_LABELS.get(r.status, r.status)
        row = (
            f"{str(r.id):<{col_widths['id']}}  "
            f"{r.name:<{col_widths['name']}}  "
            f"{r.phone or '':<{col_widths['phone']}}  "
            f"{r.session:<{col_widths['session']}}  "
            f"{label:<{col_widths['status']}}  "
            f"{r.manual_mark or '-':<{col_widths['mark']}}  "
            f"{r.notes or '-':<{col_widths['notes']}}"
        )
        click.echo(row)

    click.echo(f"\n共 {len(results)} 条结果")
    storage.close()


@main.command("view-list")
def do_view_list():
    """列出所有已保存的视图"""
    storage = _get_storage()
    views = storage.list_views()
    if not views:
        click.echo("暂无已保存的视图")
        storage.close()
        return

    for v in views:
        filters = json.loads(v["filters"])
        desc_parts = []
        if filters.get("status"):
            desc_parts.append(f"状态={filters['status']}")
        if filters.get("session"):
            desc_parts.append(f"场次={filters['session']}")
        if filters.get("mark"):
            desc_parts.append(f"标记={filters['mark']}")
        if filters.get("keyword"):
            desc_parts.append(f"关键词={filters['keyword']}")
        if filters.get("limit"):
            desc_parts.append(f"限制={filters['limit']}")
        if filters.get("sort_by"):
            desc_parts.append(f"排序={filters['sort_by']}")
        if filters.get("sort_order"):
            desc_parts.append(f"方向={filters['sort_order']}")
        desc = ", ".join(desc_parts) if desc_parts else "无筛选条件"
        click.echo(f"  {v['name']}  ({desc})")

    click.echo(f"\n共 {len(views)} 个视图")
    storage.close()


@main.command("view-delete")
@click.argument("name")
def do_view_delete(name: str):
    """删除已保存的视图"""
    storage = _get_storage()
    ok = storage.delete_view(name)
    if not ok:
        click.echo(f"错误：视图「{name}」不存在", err=True)
        storage.close()
        sys.exit(1)
    click.echo(f"[OK] 视图「{name}」已删除")
    storage.close()


@main.command("batch-mark")
@click.argument("csv_file")
def do_batch_mark(csv_file: str):
    """批量导入人工复核标记（CSV 格式：result_id, mark_text, notes）"""
    storage = _get_storage()
    rows, abs_path = _read_csv(csv_file)

    if not rows:
        click.echo("错误：CSV 文件为空", err=True)
        storage.close()
        sys.exit(1)

    headers = set(rows[0].keys())
    required = {"result_id", "mark_text"}
    missing = required - headers
    if missing:
        click.echo(f"错误：CSV 表头缺失必填列：{', '.join(sorted(missing))}", err=True)
        click.echo(f"  当前表头：{', '.join(rows[0].keys())}", err=True)
        click.echo(f"  必须包含：result_id, mark_text（notes 为可选列）", err=True)
        storage.close()
        sys.exit(1)

    prev_states, errors = batch_mark(storage, rows)

    if errors:
        click.echo(f"错误：校验未通过，共 {len(errors)} 个问题，未写入任何数据：", err=True)
        for e in errors:
            click.echo(f"  [ERR] {e['message']}", err=True)
        storage.close()
        sys.exit(1)

    click.echo(f"[OK] 成功导入 {len(prev_states)} 条人工标记")
    storage.close()


@main.command("export")
@click.option("--output", "-o", default=None, help="输出文件路径（默认从配置读取）")
@click.option("--status", "-s", default=None, help="按状态筛选")
@click.option("--session", default=None, help="按场次筛选")
@click.option("--mark", default=None, help="按人工标记筛选")
@click.option("--keyword", "-k", default=None, help="按姓名或手机号关键词筛选")
@click.option("--limit", "-n", type=int, default=None, help="限制导出条数")
@click.option("--sort-by", type=click.Choice(["id", "status"]), default=None, help="排序字段")
@click.option("--sort-order", type=click.Choice(["asc", "desc"]), default=None, help="排序方向")
@click.option("--view", "-v", default=None, help="使用已保存的视图名加载筛选条件")
@click.option("--format", "fmt", type=click.Choice(["csv", "json", "xlsx"]), default="csv", help="导出格式：csv/json/xlsx")
def do_export(output, status, session, mark, keyword, limit, sort_by, sort_order, view, fmt):
    """导出对账结果（支持 CSV/JSON/XLSX 格式，可筛选）"""
    storage = _get_storage()
    if status is not None:
        status = _validate_status(status)
    filters = _resolve_filters(
        status, session, mark, keyword, limit, sort_by, sort_order, view, None, False, storage
    )
    has_filter = any(filters.get(k) for k in ("status", "session", "mark", "keyword"))
    if has_filter:
        results = storage.query_reconcile_results(
            status=filters.get("status"),
            session=filters.get("session"),
            mark=filters.get("mark"),
            keyword=filters.get("keyword"),
            limit=filters.get("limit"),
            sort_by=filters.get("sort_by", "id"),
            sort_order=filters.get("sort_order", "asc"),
        )
    else:
        results = storage.get_all_reconcile_results()

    if not results:
        click.echo("错误：暂无对账结果，请先执行 reconcile", err=True)
        storage.close()
        sys.exit(1)

    if output is None:
        default_ext = {"csv": ".csv", "json": ".json", "xlsx": ".xlsx"}[fmt]
        base_output = config_module.get_config("export_path")
        base, ext = os.path.splitext(base_output)
        output = base + default_ext

    abs_output = os.path.abspath(output)

    status_counts = {}
    for r in results:
        status_counts[r.status] = status_counts.get(r.status, 0) + 1

    if fmt == "csv":
        os.makedirs(os.path.dirname(abs_output), exist_ok=True)
        with open(abs_output, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["ID", "姓名", "手机号", "场次", "状态", "标记", "备注"])
            for r in results:
                label = STATUS_LABELS.get(r.status, r.status)
                writer.writerow([
                    r.id,
                    r.name,
                    r.phone or "",
                    r.session,
                    label,
                    r.manual_mark or "",
                    r.notes or "",
                ])
    elif fmt == "json":
        records = []
        for r in results:
            records.append({
                "id": r.id,
                "name": r.name,
                "phone": r.phone,
                "session": r.session,
                "status": {
                    "code": r.status,
                    "label": STATUS_LABELS.get(r.status, r.status),
                },
                "manual_mark": r.manual_mark,
                "notes": r.notes,
                "references": {
                    "enroll_id": r.enroll_id,
                    "signin_id": r.signin_id,
                },
                "created_at": r.created_at,
            })
        json_data = {
            "meta": {
                "generated_at": datetime.now().isoformat(),
                "total": len(results),
                "status_summary": {
                    STATUS_LABELS.get(s, s): c for s, c in status_counts.items()
                },
            },
            "records": records,
        }
        _write_json(abs_output, json_data)
    elif fmt == "xlsx":
        headers = ["ID", "姓名", "手机号", "场次", "状态", "标记", "备注"]
        rows = []
        for r in results:
            label = STATUS_LABELS.get(r.status, r.status)
            rows.append([
                r.id,
                r.name,
                r.phone or "",
                r.session,
                label,
                r.manual_mark or "",
                r.notes or "",
            ])
        _write_xlsx(abs_output, headers, rows, sheet_name="对账结果")

    click.echo(f"[OK] 已导出 {len(results)} 条对账结果到 {abs_output}（格式：{fmt}）")
    click.echo("汇总：")
    for st in ["normal", "absent", "non_enrolled", "duplicate"]:
        cnt = status_counts.get(st, 0)
        label = STATUS_LABELS.get(st, st)
        click.echo(f"  {label}: {cnt}")

    storage.close()


def _escape_html(text: Optional[str]) -> str:
    if text is None:
        return ""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


def _build_html_report(
    title: str,
    global_stats: Dict[str, int],
    session_stats: List[Dict[str, Any]],
    results_by_session: Dict[str, List[Any]],
    generated_at: str,
) -> str:
    status_colors = {
        "normal": "#28a745",
        "absent": "#dc3545",
        "non_enrolled": "#ffc107",
        "duplicate": "#6c757d",
    }

    total = sum(global_stats.values())
    sessions_total = len(session_stats)

    def _pct(cnt: int) -> str:
        if total == 0:
            return "0.00%"
        return f"{cnt / total * 100:.2f}%"

    global_summary_rows = ""
    for status in ["normal", "absent", "non_enrolled", "duplicate"]:
        cnt = global_stats.get(status, 0)
        label = STATUS_LABELS.get(status, status)
        color = status_colors.get(status, "#333")
        global_summary_rows += f"""
        <tr>
            <td><span class="status-dot" style="background:{color}"></span>{_escape_html(label)}</td>
            <td class="num">{cnt}</td>
            <td class="num">{_pct(cnt)}</td>
        </tr>"""

    session_summary_rows = ""
    for s in session_stats:
        session_name = s["session"]
        s_total = s["total"]
        s_pct = f"{s_total / total * 100:.2f}%" if total > 0 else "0.00%"
        cells = ""
        for status in ["normal", "absent", "non_enrolled", "duplicate"]:
            cnt = s.get(status, 0)
            color = status_colors.get(status, "#333")
            pct = f"{cnt / s_total * 100:.2f}%" if s_total > 0 else "0.00%"
            cells += f'<td class="num"><span style="color:{color};font-weight:bold">{cnt}</span><br><span class="muted">{pct}</span></td>'
        session_summary_rows += f"""
        <tr>
            <td><strong>{_escape_html(session_name)}</strong></td>
            <td class="num">{s_total}</td>
            <td class="num muted">{s_pct}</td>
            {cells}
        </tr>"""

    detail_sections = ""
    for session_name, results in results_by_session.items():
        rows_html = ""
        for r in results:
            label = STATUS_LABELS.get(r.status, r.status)
            color = status_colors.get(r.status, "#333")
            rows_html += f"""
            <tr>
                <td class="num">{r.id}</td>
                <td>{_escape_html(r.name)}</td>
                <td>{_escape_html(r.phone)}</td>
                <td><span class="status-badge" style="background:{color}">{_escape_html(label)}</span></td>
                <td>{_escape_html(r.manual_mark) or "-"}</td>
                <td>{_escape_html(r.notes) or "-"}</td>
            </tr>"""
        detail_sections += f"""
        <section class="detail-section">
            <h2>场次详情：{_escape_html(session_name)} <span class="muted">（{len(results)} 人）</span></h2>
            <table class="detail-table">
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>姓名</th>
                        <th>手机号</th>
                        <th>状态</th>
                        <th>人工标记</th>
                        <th>备注</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>
        </section>"""

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{_escape_html(title)}</title>
<style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", sans-serif;
        background: #f5f7fa;
        color: #303133;
        line-height: 1.6;
        padding: 24px;
    }}
    .container {{ max-width: 1200px; margin: 0 auto; }}
    header {{
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: #fff;
        padding: 32px;
        border-radius: 12px;
        margin-bottom: 24px;
        box-shadow: 0 4px 20px rgba(102, 126, 234, 0.3);
    }}
    header h1 {{ font-size: 28px; margin-bottom: 8px; }}
    header .subtitle {{ opacity: 0.9; font-size: 14px; }}
    .card {{
        background: #fff;
        border-radius: 10px;
        padding: 24px;
        margin-bottom: 20px;
        box-shadow: 0 2px 12px rgba(0, 0, 0, 0.06);
    }}
    .card h2 {{
        font-size: 18px;
        margin-bottom: 16px;
        color: #303133;
        border-left: 4px solid #667eea;
        padding-left: 12px;
    }}
    .stats-grid {{
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
        gap: 16px;
        margin-bottom: 20px;
    }}
    .stat-card {{
        padding: 16px;
        border-radius: 8px;
        text-align: center;
        color: #fff;
    }}
    .stat-card .num {{ font-size: 28px; font-weight: bold; }}
    .stat-card .label {{ font-size: 13px; opacity: 0.9; margin-top: 4px; }}
    table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 14px;
    }}
    th, td {{
        padding: 12px 16px;
        text-align: left;
        border-bottom: 1px solid #ebeef5;
    }}
    th {{
        background: #fafafa;
        font-weight: 600;
        color: #606266;
    }}
    tr:hover td {{ background: #fafbfc; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
    th.num {{ text-align: right; }}
    .status-dot {{
        display: inline-block;
        width: 10px;
        height: 10px;
        border-radius: 50%;
        margin-right: 8px;
        vertical-align: middle;
    }}
    .status-badge {{
        display: inline-block;
        padding: 2px 10px;
        border-radius: 12px;
        color: #fff;
        font-size: 12px;
        font-weight: 500;
    }}
    .muted {{ color: #909399; font-size: 12px; }}
    .detail-section {{ margin-bottom: 32px; }}
    .detail-section h2 {{
        font-size: 16px;
        margin-bottom: 12px;
    }}
    footer {{
        text-align: center;
        color: #909399;
        font-size: 13px;
        padding: 24px;
    }}
</style>
</head>
<body>
<div class="container">
    <header>
        <h1>{_escape_html(title)}</h1>
        <div class="subtitle">生成时间：{_escape_html(generated_at)}　·　共 {sessions_total} 场次　·　{total} 条记录</div>
    </header>

    <div class="card">
        <h2>全局汇总</h2>
        <div class="stats-grid">
            <div class="stat-card" style="background: linear-gradient(135deg, #28a745, #20c997)">
                <div class="num">{global_stats.get("normal", 0)}</div>
                <div class="label">正常签到</div>
            </div>
            <div class="stat-card" style="background: linear-gradient(135deg, #dc3545, #e74c3c)">
                <div class="num">{global_stats.get("absent", 0)}</div>
                <div class="label">缺席</div>
            </div>
            <div class="stat-card" style="background: linear-gradient(135deg, #ffc107, #fd7e14)">
                <div class="num">{global_stats.get("non_enrolled", 0)}</div>
                <div class="label">非报名人员</div>
            </div>
            <div class="stat-card" style="background: linear-gradient(135deg, #6c757d, #495057)">
                <div class="num">{global_stats.get("duplicate", 0)}</div>
                <div class="label">重复扫码</div>
            </div>
        </div>
        <table>
            <thead>
                <tr>
                    <th>状态</th>
                    <th class="num">人数</th>
                    <th class="num">占比</th>
                </tr>
            </thead>
            <tbody>{global_summary_rows}
            </tbody>
        </table>
    </div>

    <div class="card">
        <h2>按场次统计</h2>
        <table>
            <thead>
                <tr>
                    <th>场次</th>
                    <th class="num">总计</th>
                    <th class="num">占比</th>
                    <th class="num">正常签到</th>
                    <th class="num">缺席</th>
                    <th class="num">非报名人员</th>
                    <th class="num">重复扫码</th>
                </tr>
            </thead>
            <tbody>{session_summary_rows}
            </tbody>
        </table>
    </div>

    <div class="card">
        {detail_sections}
    </div>

    <footer>
        由 signcheck 签到对账工具生成
    </footer>
</div>
</body>
</html>"""


@main.command("report")
@click.option("--output", "-o", default="signin_report.html", help="输出文件路径（默认 signin_report.html）")
@click.option("--session", default=None, help="按指定场次生成报告，不指定则生成全局报告")
@click.option("--format", "fmt", type=click.Choice(["html"]), default="html", help="报告格式（目前支持 html）")
def do_report(output: str, session: Optional[str], fmt: str):
    """生成签到报告（HTML 格式，含汇总统计与详细名单）"""
    storage = _get_storage()

    all_results = storage.get_all_reconcile_results()
    if not all_results:
        click.echo("错误：暂无对账结果，请先执行 reconcile", err=True)
        storage.close()
        sys.exit(1)

    all_sessions = storage.get_reconcile_sessions()

    if session is not None:
        if session not in all_sessions:
            click.echo(f"错误：场次「{session}」不存在，可用场次：{', '.join(all_sessions)}", err=True)
            storage.close()
            sys.exit(1)
        results = [r for r in all_results if r.session == session]
        target_sessions = [session]
        title = f"签到报告 - {session}"
    else:
        results = all_results
        target_sessions = all_sessions
        title = "签到报告 - 全局汇总"

    global_stats: Dict[str, int] = defaultdict(int)
    for r in results:
        global_stats[r.status] += 1

    session_status_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in results:
        session_status_counts[r.session][r.status] += 1

    session_stats: List[Dict[str, Any]] = []
    for s in target_sessions:
        counts = session_status_counts.get(s, {})
        s_total = sum(counts.values())
        row = {"session": s, "total": s_total}
        for st in ["normal", "absent", "non_enrolled", "duplicate"]:
            row[st] = counts.get(st, 0)
        session_stats.append(row)

    results_by_session: Dict[str, List[Any]] = {}
    for s in target_sessions:
        results_by_session[s] = sorted(
            [r for r in results if r.session == s],
            key=lambda r: (r.status, r.name or "")
        )

    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if fmt == "html":
        html = _build_html_report(
            title=title,
            global_stats=dict(global_stats),
            session_stats=session_stats,
            results_by_session=results_by_session,
            generated_at=generated_at,
        )
        abs_output = os.path.abspath(output)
        os.makedirs(os.path.dirname(abs_output), exist_ok=True)
        with open(abs_output, "w", encoding="utf-8") as f:
            f.write(html)

    click.echo(f"[OK] 报告已生成：{abs_output}")
    click.echo(f"  场次：{len(target_sessions)} 个　记录：{len(results)} 条")
    for st in ["normal", "absent", "non_enrolled", "duplicate"]:
        cnt = global_stats.get(st, 0)
        label = STATUS_LABELS.get(st, st)
        click.echo(f"  {label}: {cnt}")
    click.echo(f"  浏览器打开即可查看")

    storage.close()


@main.command("status")
def do_status():
    """查看当前状态"""
    storage = _get_storage()
    stats = storage.get_stats()

    click.echo(f"报名记录: {stats['enrollment_count']}")
    click.echo(f"签到记录: {stats['signin_count']}")
    click.echo(f"匹配规则: {stats['rules_count']} 条")
    click.echo(f"对账结果: {stats['result_count']} 条")

    if stats["status_counts"]:
        for status in ["normal", "absent", "non_enrolled", "duplicate"]:
            cnt = stats["status_counts"].get(status, 0)
            if cnt:
                label = STATUS_LABELS.get(status, status)
                click.echo(f"  - {label}: {cnt}")

    click.echo(f"撤销历史: {stats['undo_count']} 步")
    click.echo(f"导入错误: {stats['error_count']} 条")

    if stats['error_count'] > 0:
        errors = storage.get_all_import_errors()
        migration_errors = [e for e in errors if e.source_type.startswith("migration_")]
        if migration_errors:
            enroll_dropped = sum(1 for e in migration_errors if e.source_type == "migration_enrollments")
            signin_dropped = sum(1 for e in migration_errors if e.source_type == "migration_signins")
            if enroll_dropped:
                click.echo(f"  （迁移时去重丢弃报名重复记录 {enroll_dropped} 条，详见 errors 命令）")
            if signin_dropped:
                click.echo(f"  （迁移时去重丢弃签到重复记录 {signin_dropped} 条，详见 errors 命令）")

    storage.close()


@main.command("errors")
def show_errors():
    """查看导入错误"""
    storage = _get_storage()
    errors = storage.get_all_import_errors()

    if not errors:
        click.echo("暂无导入错误记录")
        storage.close()
        return

    for e in errors:
        click.echo(f"[{e.source_type}] {e.error_message}")
        if e.raw_data:
            click.echo(f"  原始数据: {e.raw_data}")

    click.echo(f"\n共 {len(errors)} 条错误")
    storage.close()


@main.command("reset")
@click.confirmation_option(prompt="确定要清空所有数据吗？")
def do_reset():
    """清空所有数据（需确认）"""
    db_dir = os.path.join(os.getcwd(), ".signcheck")
    db_path = os.path.join(db_dir, "signcheck.db")
    if os.path.exists(db_path):
        os.remove(db_path)
        click.echo("[OK] 已清空所有数据")
    else:
        click.echo("当前无数据")


@main.command("backup")
@click.option("--output-dir", "-d", default=".", help="备份文件输出目录（默认当前目录）")
def do_backup(output_dir: str):
    """备份全部数据为时间戳命名的 .zip 快照"""
    storage = _get_storage()
    snapshot = storage.export_all()
    storage.close()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_name = f"signcheck_backup_{timestamp}.zip"
    abs_output_dir = os.path.abspath(output_dir)
    os.makedirs(abs_output_dir, exist_ok=True)
    abs_output = os.path.join(abs_output_dir, backup_name)

    tmp_dir = tempfile.mkdtemp(prefix="signcheck_backup_")
    try:
        meta_file = os.path.join(tmp_dir, "metadata.json")
        meta = snapshot.get("_meta", {})
        meta["filename"] = backup_name
        _write_json(meta_file, meta)

        for table_name in Storage.BACKUP_TABLES:
            rows = snapshot.get(table_name, [])
            if rows:
                table_file = os.path.join(tmp_dir, f"{table_name}.json")
                _write_json(table_file, rows)

        with zipfile.ZipFile(abs_output, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.write(meta_file, "metadata.json")
            for table_name in Storage.BACKUP_TABLES:
                table_file = os.path.join(tmp_dir, f"{table_name}.json")
                if os.path.exists(table_file):
                    zf.write(table_file, f"{table_name}.json")
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    click.echo(f"[OK] 备份已生成：{abs_output}")
    summary = snapshot.get("_meta", {})
    for table in Storage.BACKUP_TABLES:
        cnt = len(snapshot.get(table, []))
        if cnt > 0:
            click.echo(f"  {table}: {cnt} 条")
    click.echo(f"  导出时间：{summary.get('exported_at', 'N/A')}")


@main.command("restore")
@click.argument("backup_file")
@click.option("--force", is_flag=True, default=False, help="跳过冲突确认，强制覆盖还原")
def do_restore(backup_file: str, force: bool):
    """从备份 .zip 还原数据，还原前检查冲突"""
    abs_backup = os.path.abspath(backup_file)
    if not os.path.exists(abs_backup):
        click.echo(f"错误：备份文件不存在 {abs_backup}", err=True)
        sys.exit(1)
    if not zipfile.is_zipfile(abs_backup):
        click.echo(f"错误：{abs_backup} 不是有效的 zip 文件", err=True)
        sys.exit(1)

    tmp_dir = tempfile.mkdtemp(prefix="signcheck_restore_")
    try:
        with zipfile.ZipFile(abs_backup, "r") as zf:
            zf.extractall(tmp_dir)

        meta_file = os.path.join(tmp_dir, "metadata.json")
        if not os.path.exists(meta_file):
            click.echo("错误：备份文件缺少 metadata.json", err=True)
            sys.exit(1)
        with open(meta_file, "r", encoding="utf-8") as f:
            meta = json.load(f)

        snapshot: Dict[str, Any] = {"_meta": meta}
        for table_name in Storage.BACKUP_TABLES:
            table_file = os.path.join(tmp_dir, f"{table_name}.json")
            if os.path.exists(table_file):
                with open(table_file, "r", encoding="utf-8") as f:
                    snapshot[table_name] = json.load(f)
            else:
                snapshot[table_name] = []
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)

    storage = _get_storage()
    try:
        current = storage.get_conflict_summary()
        backup_sessions = set()
        for s in snapshot.get("sessions", []):
            if s.get("name"):
                backup_sessions.add(s["name"])
        for e in snapshot.get("enrollments", []):
            if e.get("session"):
                backup_sessions.add(e["session"])
        for s in snapshot.get("signins", []):
            if s.get("session"):
                backup_sessions.add(s["session"])
        for r in snapshot.get("reconcile_results", []):
            if r.get("session"):
                backup_sessions.add(r["session"])

        current_sessions = set(current.get("sessions", []))
        session_conflicts = sorted(backup_sessions & current_sessions)

        table_conflicts = []
        for table in Storage.BACKUP_TABLES:
            cur_cnt = current.get("table_counts", {}).get(table, 0)
            bak_cnt = len(snapshot.get(table, []))
            if cur_cnt > 0 and bak_cnt > 0:
                table_conflicts.append((table, cur_cnt, bak_cnt))

        has_conflict = bool(session_conflicts or table_conflicts)

        if has_conflict and not force:
            click.echo("检测到冲突，请确认是否继续还原：")
            click.echo()
            if session_conflicts:
                click.echo(f"场次名冲突（{len(session_conflicts)} 个）：")
                for s in session_conflicts:
                    click.echo(f"  - {s}")
                click.echo()
            if table_conflicts:
                click.echo("现有数据与备份数据重叠的表：")
                for table, cur_cnt, bak_cnt in table_conflicts:
                    click.echo(f"  - {table}: 当前 {cur_cnt} 条 / 备份 {bak_cnt} 条")
                click.echo()
            if not click.confirm("确认用备份数据覆盖当前全部数据？此操作不可撤销"):
                click.echo("已取消还原")
                storage.close()
                return

        storage.import_snapshot(snapshot, overwrite=True)

        click.echo(f"[OK] 已从备份还原：{abs_backup}")
        for table in Storage.BACKUP_TABLES:
            cnt = len(snapshot.get(table, []))
            if cnt > 0:
                click.echo(f"  {table}: 还原 {cnt} 条")
    finally:
        storage.close()


def _calc_percentage(count: int, total: int) -> str:
    if total == 0:
        return "0.00%"
    return f"{count / total * 100:.2f}%"


def _build_stats_rows(results: List[Dict[str, Any]], sessions: List[str]) -> List[Dict[str, Any]]:
    session_status_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in results:
        session_status_counts[r["session"]][r["status"]] = r["count"]

    rows = []
    for session in sessions:
        counts = session_status_counts.get(session, {})
        total = sum(counts.values())
        row = {
            "session": session,
            "total": total,
        }
        for status in ["normal", "absent", "non_enrolled", "duplicate"]:
            cnt = counts.get(status, 0)
            row[status] = cnt
            row[f"{status}_pct"] = _calc_percentage(cnt, total)
        rows.append(row)
    return rows


def _print_stats_table(rows: List[Dict[str, Any]], show_global: bool = True):
    all_statuses = ["normal", "absent", "non_enrolled", "duplicate"]

    if show_global and len(rows) > 1:
        global_row = {
            "session": "合计",
            "total": sum(r["total"] for r in rows),
        }
        for status in all_statuses:
            global_row[status] = sum(r[status] for r in rows)
        total_all = global_row["total"]
        for status in all_statuses:
            global_row[f"{status}_pct"] = _calc_percentage(global_row[status], total_all)
        display_rows = rows + [global_row]
    else:
        display_rows = rows

    col_widths = {
        "session": max(6, max(len(str(r["session"])) for r in display_rows)),
        "total": max(6, max(len(str(r["total"])) for r in display_rows)),
    }
    for status in all_statuses:
        label = STATUS_LABELS[status]
        max_cnt_len = max(len(str(r[status])) for r in display_rows)
        max_pct_len = max(len(r[f"{status}_pct"]) for r in display_rows)
        col_widths[status] = max(len(label), max_cnt_len + max_pct_len + 3)

    header = f"{'场次':<{col_widths['session']}}  {'总计':<{col_widths['total']}}"
    for status in all_statuses:
        label = STATUS_LABELS[status]
        header += f"  {label:<{col_widths[status]}}"
    click.echo(header)
    click.echo("-" * len(header))

    for r in display_rows:
        line = f"{str(r['session']):<{col_widths['session']}}  {str(r['total']):<{col_widths['total']}}"
        for status in all_statuses:
            cnt = r[status]
            pct = r[f"{status}_pct"]
            cell = f"{cnt} ({pct})"
            line += f"  {cell:<{col_widths[status]}}"
        click.echo(line)


@main.command("stats")
@click.option("--session", default=None, help="按指定场次查看统计")
@click.option("--export", "-o", default=None, help="导出统计结果到文件")
@click.option("--format", "fmt", type=click.Choice(["csv", "json", "xlsx"]), default="csv", help="导出格式：csv/json/xlsx")
def do_stats(session: Optional[str], export: Optional[str], fmt: str):
    """查看对账结果统计汇总"""
    storage = _get_storage()

    all_results = storage.count_reconcile_results_by_session_and_status()
    all_sessions = storage.get_reconcile_sessions()

    if not all_results:
        click.echo("暂无对账结果，请先执行 reconcile")
        storage.close()
        return

    if session is not None:
        if session not in all_sessions:
            click.echo(f"错误：场次「{session}」不存在，可用场次：{', '.join(all_sessions)}", err=True)
            storage.close()
            sys.exit(1)
        results = [r for r in all_results if r["session"] == session]
        sessions = [session]
    else:
        results = all_results
        sessions = all_sessions

    stats_rows = _build_stats_rows(results, sessions)

    if export is None:
        _print_stats_table(stats_rows, show_global=(session is None and len(sessions) > 1))
        total_all = sum(r["total"] for r in stats_rows)
        click.echo(f"\n共 {len(stats_rows)} 场，{total_all} 条记录")
    else:
        abs_output = os.path.abspath(export)
        headers = ["场次", "总计", "正常签到", "正常签到占比", "缺席", "缺席占比", "非报名人员", "非报名人员占比", "重复扫码", "重复扫码占比"]
        display_rows = []
        for r in stats_rows:
            display_rows.append([
                r["session"],
                r["total"],
                r["normal"],
                r["normal_pct"],
                r["absent"],
                r["absent_pct"],
                r["non_enrolled"],
                r["non_enrolled_pct"],
                r["duplicate"],
                r["duplicate_pct"],
            ])
        if session is None and len(stats_rows) > 1:
            total_all = sum(r["total"] for r in stats_rows)
            display_rows.append([
                "合计",
                total_all,
                sum(r["normal"] for r in stats_rows),
                _calc_percentage(sum(r["normal"] for r in stats_rows), total_all),
                sum(r["absent"] for r in stats_rows),
                _calc_percentage(sum(r["absent"] for r in stats_rows), total_all),
                sum(r["non_enrolled"] for r in stats_rows),
                _calc_percentage(sum(r["non_enrolled"] for r in stats_rows), total_all),
                sum(r["duplicate"] for r in stats_rows),
                _calc_percentage(sum(r["duplicate"] for r in stats_rows), total_all),
            ])

        if fmt == "csv":
            os.makedirs(os.path.dirname(abs_output), exist_ok=True)
            with open(abs_output, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for row in display_rows:
                    writer.writerow(row)
        elif fmt == "json":
            session_stats = []
            for r in stats_rows:
                session_stats.append({
                    "session": r["session"],
                    "total": r["total"],
                    "status": {
                        STATUS_LABELS[s]: {
                            "count": r[s],
                            "percentage": r[f"{s}_pct"],
                        } for s in ["normal", "absent", "non_enrolled", "duplicate"]
                    },
                })
            json_data = {
                "meta": {
                    "generated_at": datetime.now().isoformat(),
                    "sessions_count": len(stats_rows),
                    "total_records": sum(r["total"] for r in stats_rows),
                },
                "sessions": session_stats,
            }
            if session is None and len(stats_rows) > 1:
                total_all = sum(r["total"] for r in stats_rows)
                json_data["global"] = {
                    "total": total_all,
                    "status": {
                        STATUS_LABELS[s]: {
                            "count": sum(r[s] for r in stats_rows),
                            "percentage": _calc_percentage(sum(r[s] for r in stats_rows), total_all),
                        } for s in ["normal", "absent", "non_enrolled", "duplicate"]
                    },
                }
            _write_json(abs_output, json_data)
        elif fmt == "xlsx":
            _write_xlsx(abs_output, headers, display_rows, sheet_name="场次统计")

        click.echo(f"[OK] 已导出 {len(stats_rows)} 场统计到 {abs_output}（格式：{fmt}）")

    storage.close()


@main.group("session")
def session_group():
    """管理场次生命周期"""
    pass


@session_group.command("create")
@click.argument("name")
@click.option("--start", "start_time", default=None, help="场次开始时间")
@click.option("--end", "end_time", default=None, help="场次结束时间")
@click.option("--desc", "description", default=None, help="场次描述")
def session_create(name: str, start_time: Optional[str], end_time: Optional[str], description: Optional[str]):
    """创建新场次"""
    storage = _get_storage()
    session_id = storage.create_session(name, start_time, end_time, description)
    if session_id is None:
        click.echo(f"错误：场次「{name}」已存在", err=True)
        storage.close()
        sys.exit(1)
    click.echo(f"[OK] 已创建场次「{name}」（ID: {session_id}）")
    storage.close()


@session_group.command("close")
@click.argument("name")
def session_close(name: str):
    """关闭场次，拒绝新导入和对账操作"""
    storage = _get_storage()
    if storage.get_session(name) is None:
        click.echo(f"错误：场次「{name}」不存在", err=True)
        storage.close()
        sys.exit(1)
    ok = storage.close_session(name)
    if not ok:
        click.echo(f"错误：场次「{name}」已处于关闭状态", err=True)
        storage.close()
        sys.exit(1)
    click.echo(f"[OK] 场次「{name}」已关闭")
    storage.close()


@session_group.command("list")
def session_list():
    """列出所有场次及状态"""
    storage = _get_storage()
    sessions = storage.get_all_sessions()
    if not sessions:
        click.echo("暂无场次，请使用 session create 创建")
        storage.close()
        return

    col_widths = {
        "name": max(6, max(len(s.name) for s in sessions)),
        "status": max(8, max(len("已关闭" if s.status == "closed" else "进行中") for s in sessions)),
        "start": max(8, max(len(s.start_time or "-") for s in sessions)),
        "end": max(8, max(len(s.end_time or "-") for s in sessions)),
        "created": max(8, max(len(s.created_at or "-") for s in sessions)),
    }

    header = (
        f"{'场次名':<{col_widths['name']}}  "
        f"{'状态':<{col_widths['status']}}  "
        f"{'开始时间':<{col_widths['start']}}  "
        f"{'结束时间':<{col_widths['end']}}  "
        f"{'创建时间':<{col_widths['created']}}"
    )
    click.echo(header)
    click.echo("-" * len(header))

    for s in sessions:
        status_label = "已关闭" if s.status == "closed" else "进行中"
        row = (
            f"{s.name:<{col_widths['name']}}  "
            f"{status_label:<{col_widths['status']}}  "
            f"{s.start_time or '-':<{col_widths['start']}}  "
            f"{s.end_time or '-':<{col_widths['end']}}  "
            f"{s.created_at or '-':<{col_widths['created']}}"
        )
        click.echo(row)

    click.echo(f"\n共 {len(sessions)} 个场次")
    storage.close()


@main.group("config")
def config_group():
    """管理 CLI 配置偏好"""
    pass


@config_group.command("show")
def config_show():
    """显示当前配置"""
    cfg = config_module.load_config()
    click.echo("当前配置：")
    for key in sorted(cfg.keys()):
        click.echo(f"  {key}: {cfg[key]}")


@config_group.command("set")
@click.argument("key")
@click.argument("value")
def config_set(key: str, value: str):
    """设置单项配置"""
    try:
        cfg = config_module.set_config(key, value)
        click.echo(f"[OK] 已设置 {key} = {cfg[key]}")
    except ValueError as e:
        click.echo(f"错误：{e}", err=True)
        sys.exit(1)


@config_group.command("reset")
@click.confirmation_option(prompt="确定要恢复出厂默认配置吗？")
def config_reset():
    """恢复出厂默认配置"""
    cfg = config_module.reset_config()
    click.echo("[OK] 已恢复默认配置")
    click.echo("当前配置：")
    for key in sorted(cfg.keys()):
        click.echo(f"  {key}: {cfg[key]}")


if __name__ == "__main__":
    main()
