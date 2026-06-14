import csv
import io
import json
import os
from collections import defaultdict
from datetime import datetime
from typing import Any, Dict, List, Optional

from .constants import (
    CSV_EXPORT_HEADERS,
    STATUS_COLORS,
    STATUS_LABELS,
    STATUS_ORDER,
    STATS_EXPORT_HEADERS,
)


def result_to_dict(r) -> dict:
    return {
        "id": r.id,
        "enroll_id": r.enroll_id,
        "signin_id": r.signin_id,
        "name": r.name,
        "phone": r.phone,
        "session": r.session,
        "status": {
            "code": r.status,
            "label": STATUS_LABELS.get(r.status, r.status),
        },
        "manual_mark": r.manual_mark,
        "notes": r.notes,
        "created_at": r.created_at,
    }


def compute_status_counts(results) -> Dict[str, int]:
    counts: Dict[str, int] = defaultdict(int)
    for r in results:
        counts[r.status] += 1
    return dict(counts)


def format_csv_content(results) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(CSV_EXPORT_HEADERS)
    for r in results:
        label = STATUS_LABELS.get(r.status, r.status)
        writer.writerow([
            r.id, r.name, r.phone or "", r.session,
            label, r.manual_mark or "", r.notes or "",
        ])
    return buf.getvalue()


def write_csv_file(file_path: str, results) -> str:
    abs_path = os.path.abspath(file_path)
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)
    with open(abs_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(CSV_EXPORT_HEADERS)
        for r in results:
            label = STATUS_LABELS.get(r.status, r.status)
            writer.writerow([
                r.id, r.name, r.phone or "", r.session,
                label, r.manual_mark or "", r.notes or "",
            ])
    return abs_path


def build_json_data(results, status_counts: Optional[Dict[str, int]] = None) -> dict:
    if status_counts is None:
        status_counts = compute_status_counts(results)
    records = []
    for r in results:
        records.append({
            "id": r.id, "name": r.name, "phone": r.phone,
            "session": r.session,
            "status": {"code": r.status, "label": STATUS_LABELS.get(r.status, r.status)},
            "manual_mark": r.manual_mark, "notes": r.notes,
            "references": {"enroll_id": r.enroll_id, "signin_id": r.signin_id},
            "created_at": r.created_at,
        })
    return {
        "meta": {
            "generated_at": datetime.now().isoformat(),
            "total": len(results),
            "status_summary": {STATUS_LABELS.get(s, s): c for s, c in status_counts.items()},
        },
        "records": records,
    }


def write_json_file(file_path: str, data: Any) -> str:
    abs_path = os.path.abspath(file_path)
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)
    with open(abs_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return abs_path


def write_xlsx_file(file_path: str, headers: List[str], rows: List[List[Any]], sheet_name: str = "Sheet1") -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl 未安装，请执行：pip install openpyxl>=3.1.0")

    abs_path = os.path.abspath(file_path)
    os.makedirs(os.path.dirname(abs_path), exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font
        cell.alignment = center_align

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    ws.freeze_panes = "A2"

    for col_idx in range(1, len(headers) + 1):
        max_len = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.value is not None:
                    cell_len = len(str(cell.value))
                    if cell_len > max_len:
                        max_len = cell_len
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    wb.save(abs_path)
    return abs_path


def build_results_xlsx_rows(results) -> tuple:
    headers = CSV_EXPORT_HEADERS
    rows = []
    for r in results:
        label = STATUS_LABELS.get(r.status, r.status)
        rows.append([
            r.id, r.name, r.phone or "", r.session,
            label, r.manual_mark or "", r.notes or "",
        ])
    return headers, rows


def escape_html(text: Optional[str]) -> str:
    if text is None:
        return ""
    return (
        str(text)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&#39;")
    )


def build_html_report(
    title: str,
    global_stats: Dict[str, int],
    session_stats: List[Dict[str, Any]],
    results_by_session: Dict[str, List[Any]],
    generated_at: str,
) -> str:
    total = sum(global_stats.values())
    sessions_total = len(session_stats)

    def _pct(cnt: int) -> str:
        if total == 0:
            return "0.00%"
        return f"{cnt / total * 100:.2f}%"

    global_summary_rows = ""
    for status in STATUS_ORDER:
        cnt = global_stats.get(status, 0)
        label = STATUS_LABELS.get(status, status)
        color = STATUS_COLORS.get(status, "#333")
        global_summary_rows += f"""
        <tr>
            <td><span class="status-dot" style="background:{color}"></span>{escape_html(label)}</td>
            <td class="num">{cnt}</td>
            <td class="num">{_pct(cnt)}</td>
        </tr>"""

    session_summary_rows = ""
    for s in session_stats:
        session_name = s["session"]
        s_total = s["total"]
        s_pct = f"{s_total / total * 100:.2f}%" if total > 0 else "0.00%"
        cells = ""
        for status in STATUS_ORDER:
            cnt = s.get(status, 0)
            color = STATUS_COLORS.get(status, "#333")
            pct = f"{cnt / s_total * 100:.2f}%" if s_total > 0 else "0.00%"
            cells += f'<td class="num"><span style="color:{color};font-weight:bold">{cnt}</span><br><span class="muted">{pct}</span></td>'
        session_summary_rows += f"""
        <tr>
            <td><strong>{escape_html(session_name)}</strong></td>
            <td class="num">{s_total}</td>
            <td class="num muted">{s_pct}</td>
            {cells}
        </tr>"""

    detail_sections = ""
    for session_name, results in results_by_session.items():
        rows_html = ""
        for r in results:
            label = STATUS_LABELS.get(r.status, r.status)
            color = STATUS_COLORS.get(r.status, "#333")
            rows_html += f"""
            <tr>
                <td class="num">{r.id}</td>
                <td>{escape_html(r.name)}</td>
                <td>{escape_html(r.phone)}</td>
                <td><span class="status-badge" style="background:{color}">{escape_html(label)}</span></td>
                <td>{escape_html(r.manual_mark) or "-"}</td>
                <td>{escape_html(r.notes) or "-"}</td>
            </tr>"""
        detail_sections += f"""
        <section class="detail-section">
            <h2>场次详情：{escape_html(session_name)} <span class="muted">（{len(results)} 人）</span></h2>
            <table class="detail-table">
                <thead>
                    <tr>
                        <th>ID</th>
                        <th>姓名</th>
                        <th>手机号</th>
                        <th>状态</th>
                        <th>人工标记</th>
                        <th>备注</th>
                    </tr>
                </thead>
                <tbody>
                    {rows_html}
                </tbody>
            </table>
        </section>"""

    stat_card_colors = {
        "normal": "background: linear-gradient(135deg, #28a745, #20c997)",
        "absent": "background: linear-gradient(135deg, #dc3545, #e74c3c)",
        "non_enrolled": "background: linear-gradient(135deg, #ffc107, #fd7e14)",
        "duplicate": "background: linear-gradient(135deg, #6c757d, #495057)",
    }
    stat_card_labels = {
        "normal": "正常签到",
        "absent": "缺席",
        "non_enrolled": "非报名人员",
        "duplicate": "重复扫码",
    }

    stat_cards = ""
    for status in STATUS_ORDER:
        cnt = global_stats.get(status, 0)
        bg = stat_card_colors.get(status, "")
        lbl = stat_card_labels.get(status, "")
        stat_cards += f"""
            <div class="stat-card" style="{bg}">
                <div class="num">{cnt}</div>
                <div class="label">{lbl}</div>
            </div>"""

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{escape_html(title)}</title>
<style>
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
        font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", sans-serif;
        background: #f5f7fa;
        color: #303133;
        line-height: 1.6;
        padding: 24px;
    }}
    .container {{ max-width: 1200px; margin: 0 auto; }}
    header {{
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: #fff;
        padding: 32px;
        border-radius: 12px;
        margin-bottom: 24px;
        box-shadow: 0 4px 20px rgba(102, 126, 234, 0.3);
    }}
    header h1 {{ font-size: 28px; margin-bottom: 8px; }}
    header .subtitle {{ opacity: 0.9; font-size: 14px; }}
    .card {{
        background: #fff;
        border-radius: 10px;
        padding: 24px;
        margin-bottom: 20px;
        box-shadow: 0 2px 12px rgba(0, 0, 0, 0.06);
    }}
    .card h2 {{
        font-size: 18px;
        margin-bottom: 16px;
        color: #303133;
        border-left: 4px solid #667eea;
        padding-left: 12px;
    }}
    .stats-grid {{
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(160px, 1fr));
        gap: 16px;
        margin-bottom: 20px;
    }}
    .stat-card {{
        padding: 16px;
        border-radius: 8px;
        text-align: center;
        color: #fff;
    }}
    .stat-card .num {{ font-size: 28px; font-weight: bold; }}
    .stat-card .label {{ font-size: 13px; opacity: 0.9; margin-top: 4px; }}
    table {{
        width: 100%;
        border-collapse: collapse;
        font-size: 14px;
    }}
    th, td {{
        padding: 12px 16px;
        text-align: left;
        border-bottom: 1px solid #ebeef5;
    }}
    th {{
        background: #fafafa;
        font-weight: 600;
        color: #606266;
    }}
    tr:hover td {{ background: #fafbfc; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
    th.num {{ text-align: right; }}
    .status-dot {{
        display: inline-block;
        width: 10px;
        height: 10px;
        border-radius: 50%;
        margin-right: 8px;
        vertical-align: middle;
    }}
    .status-badge {{
        display: inline-block;
        padding: 2px 10px;
        border-radius: 12px;
        color: #fff;
        font-size: 12px;
        font-weight: 500;
    }}
    .muted {{ color: #909399; font-size: 12px; }}
    .detail-section {{ margin-bottom: 32px; }}
    .detail-section h2 {{
        font-size: 16px;
        margin-bottom: 12px;
    }}
    footer {{
        text-align: center;
        color: #909399;
        font-size: 13px;
        padding: 24px;
    }}
</style>
</head>
<body>
<div class="container">
    <header>
        <h1>{escape_html(title)}</h1>
        <div class="subtitle">生成时间：{escape_html(generated_at)}　·　共 {sessions_total} 场次　·　{total} 条记录</div>
    </header>

    <div class="card">
        <h2>全局汇总</h2>
        <div class="stats-grid">
            {stat_cards}
        </div>
        <table>
            <thead>
                <tr>
                    <th>状态</th>
                    <th class="num">人数</th>
                    <th class="num">占比</th>
                </tr>
            </thead>
            <tbody>{global_summary_rows}
            </tbody>
        </table>
    </div>

    <div class="card">
        <h2>按场次统计</h2>
        <table>
            <thead>
                <tr>
                    <th>场次</th>
                    <th class="num">总计</th>
                    <th class="num">占比</th>
                    <th class="num">正常签到</th>
                    <th class="num">缺席</th>
                    <th class="num">非报名人员</th>
                    <th class="num">重复扫码</th>
                </tr>
            </thead>
            <tbody>{session_summary_rows}
            </tbody>
        </table>
    </div>

    <div class="card">
        {detail_sections}
    </div>

    <footer>
        由 signcheck 签到对账工具生成
    </footer>
</div>
</body>
</html>"""


def calc_percentage(count: int, total: int) -> str:
    if total == 0:
        return "0.00%"
    return f"{count / total * 100:.2f}%"


def build_stats_rows(results: List[Dict[str, Any]], sessions: List[str]) -> List[Dict[str, Any]]:
    session_status_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in results:
        session_status_counts[r["session"]][r["status"]] = r["count"]

    rows = []
    for session in sessions:
        counts = session_status_counts.get(session, {})
        total = sum(counts.values())
        row: Dict[str, Any] = {"session": session, "total": total}
        for status in STATUS_ORDER:
            cnt = counts.get(status, 0)
            row[status] = cnt
            row[f"{status}_pct"] = calc_percentage(cnt, total)
        rows.append(row)
    return rows


def build_session_stats_for_report(results, target_sessions: List[str]) -> tuple:
    global_stats: Dict[str, int] = defaultdict(int)
    for r in results:
        global_stats[r.status] += 1

    session_status_counts: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
    for r in results:
        session_status_counts[r.session][r.status] += 1

    session_stats: List[Dict[str, Any]] = []
    for s in target_sessions:
        counts = session_status_counts.get(s, {})
        s_total = sum(counts.values())
        row: Dict[str, Any] = {"session": s, "total": s_total}
        for st in STATUS_ORDER:
            row[st] = counts.get(st, 0)
        session_stats.append(row)

    results_by_session: Dict[str, List[Any]] = {}
    for s in target_sessions:
        results_by_session[s] = sorted(
            [r for r in results if r.session == s],
            key=lambda r: (r.status, r.name or ""),
        )

    return dict(global_stats), session_stats, results_by_session


def build_session_detail(storage, session_name: str, all_results) -> dict:
    from .constants import STATUS_LABELS, STATUS_ORDER
    session_results = [r for r in all_results if r.session == session_name]
    total = len(session_results)
    status_counts: Dict[str, int] = defaultdict(int)
    for r in session_results:
        status_counts[r.status] += 1
    pct = lambda c: f"{c / total * 100:.2f}%" if total > 0 else "0.00%"
    return {
        "session": session_name,
        "total": total,
        "status_breakdown": {
            s: {
                "count": status_counts.get(s, 0),
                "label": STATUS_LABELS.get(s, s),
                "percentage": pct(status_counts.get(s, 0)),
            }
            for s in STATUS_ORDER
        },
    }
